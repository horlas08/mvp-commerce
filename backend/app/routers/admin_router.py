import os
import shutil
import uuid
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile, BackgroundTasks
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models.user import User
from app.models.product import Product
from app.models.order import Order, OrderItem, OrderStatus, PaymentStatus
from app.models.category import Category
from app.models.coupon import Coupon
from app.models.banner import Banner
from app.models.location import State, City
from app.models.refund import RefundRequest, RefundStatus
from app.models.notification import Notification
from app.models.support import SupportTicket, SupportMessage
from app.models.wallet_transaction import WalletTransaction, WalletTransactionType
from app.auth.dependencies import get_current_user

router = APIRouter(prefix="/admin", tags=["Admin"])


# ── Admin guard dependency ──────────────────────────────────────────────────

async def get_admin_user(user: User = Depends(get_current_user)) -> User:
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user


@router.post("/upload-image")
async def upload_image(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    # Validate extension
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in [".jpg", ".jpeg", ".png", ".gif", ".webp"]:
        raise HTTPException(status_code=400, detail="Only .jpg, .jpeg, .png, .gif, .webp files are allowed")

    # Create directory path
    static_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "uploads", "images")
    os.makedirs(static_dir, exist_ok=True)

    # Save to unique filename
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = os.path.join(static_dir, unique_filename)

    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    image_url = f"/static/uploads/images/{unique_filename}"
    return {"image_url": image_url}


# ── Stats ───────────────────────────────────────────────────────────────────

@router.get("/stats")
async def get_stats(db: AsyncSession = Depends(get_db), _: User = Depends(get_admin_user)):
    """Overall dashboard stats."""
    total_users = (await db.execute(select(func.count(User.id)))).scalar_one()
    total_products = (await db.execute(select(func.count(Product.id)))).scalar_one()
    total_orders = (await db.execute(select(func.count(Order.id)))).scalar_one()
    total_revenue = (await db.execute(select(func.coalesce(func.sum(Order.total), 0.0)))).scalar_one()
    pending_orders = (await db.execute(
        select(func.count(Order.id)).where(Order.status == OrderStatus.PENDING)
    )).scalar_one()
    active_products = (await db.execute(
        select(func.count(Product.id)).where(Product.is_active == True)
    )).scalar_one()
    pending_payments = (await db.execute(
        select(func.count(Order.id)).where(Order.payment_status == PaymentStatus.PENDING_APPROVAL)
    )).scalar_one()
    pending_refunds = (await db.execute(
        select(func.count(RefundRequest.id)).where(RefundRequest.status == RefundStatus.PENDING)
    )).scalar_one()

    return {
        "total_users": total_users,
        "total_products": total_products,
        "total_orders": total_orders,
        "total_revenue": float(total_revenue),
        "pending_orders": pending_orders,
        "active_products": active_products,
        "pending_payments": pending_payments,
        "pending_refunds": pending_refunds,
    }


# ── Users ───────────────────────────────────────────────────────────────────

@router.get("/users")
async def list_users(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    search: Optional[str] = None,
    role: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    query = select(User)
    if search:
        query = query.where(
            (User.email.ilike(f"%{search}%")) | (User.name.ilike(f"%{search}%"))
        )
    if role:
        query = query.where(User.role == role)
    query = query.order_by(User.created_at.desc()).offset((page - 1) * limit).limit(limit)
    result = await db.execute(query)
    users = result.scalars().all()
    count_query = select(func.count(User.id))
    if search:
        count_query = count_query.where(
            (User.email.ilike(f"%{search}%")) | (User.name.ilike(f"%{search}%"))
        )
    if role:
        count_query = count_query.where(User.role == role)
    total = (await db.execute(count_query)).scalar_one()
    return {"users": [u.to_dict() for u in users], "total": total, "page": page, "limit": limit}


class UpdateUserRequest(BaseModel):
    role: Optional[str] = None
    is_active: Optional[bool] = None
    credit_balance: Optional[float] = None


@router.patch("/users/{user_id}")
async def update_user(
    user_id: str,
    req: UpdateUserRequest,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_admin_user),
):
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if req.role is not None:
        user.role = req.role
    if req.is_active is not None:
        user.is_active = req.is_active
    if req.credit_balance is not None:
        user.credit_balance = req.credit_balance
    await db.commit()
    return user.to_dict()


@router.delete("/users/{user_id}")
async def delete_user(
    user_id: str,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_admin_user),
):
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    await db.delete(user)
    await db.commit()
    return {"message": "User deleted"}


# ── Products ─────────────────────────────────────────────────────────────────

@router.get("/products")
async def list_products_admin(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    search: Optional[str] = None,
    category_id: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    query = select(Product)
    if search:
        query = query.where(
            (Product.title_en.ilike(f"%{search}%")) | (Product.title_ar.ilike(f"%{search}%"))
        )
    if category_id:
        query = query.where(Product.category_id == category_id)
    query = query.order_by(Product.created_at.desc()).offset((page - 1) * limit).limit(limit)
    result = await db.execute(query)
    products = result.scalars().all()

    count_query = select(func.count(Product.id))
    if search:
        count_query = count_query.where(
            (Product.title_en.ilike(f"%{search}%")) | (Product.title_ar.ilike(f"%{search}%"))
        )
    if category_id:
        count_query = count_query.where(Product.category_id == category_id)
    total = (await db.execute(count_query)).scalar_one()

    return {"products": [p.to_dict() for p in products], "total": total, "page": page, "limit": limit}


class CreateProductRequest(BaseModel):
    title_en: str
    title_ar: str
    description_en: Optional[str] = None
    description_ar: Optional[str] = None
    price: float
    discount_price: Optional[float] = None
    category_id: Optional[str] = None
    stock: int = 0
    images: Optional[list] = None


@router.post("/products")
async def create_product(
    req: CreateProductRequest,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    product = Product(**req.model_dump())
    db.add(product)
    await db.commit()
    await db.refresh(product)
    return product.to_dict()


class UpdateProductRequest(BaseModel):
    title_en: Optional[str] = None
    title_ar: Optional[str] = None
    description_en: Optional[str] = None
    description_ar: Optional[str] = None
    price: Optional[float] = None
    discount_price: Optional[float] = None
    category_id: Optional[str] = None
    stock: Optional[int] = None
    is_active: Optional[bool] = None
    images: Optional[list] = None


@router.patch("/products/{product_id}")
async def update_product(
    product_id: str,
    req: UpdateProductRequest,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    result = await db.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    for field, value in req.model_dump(exclude_none=True).items():
        setattr(product, field, value)
    await db.commit()
    return product.to_dict()


@router.delete("/products/{product_id}")
async def delete_product(
    product_id: str,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    result = await db.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    await db.delete(product)
    await db.commit()
    return {"message": "Product deleted"}


# ── Orders ───────────────────────────────────────────────────────────────────

@router.get("/orders")
async def list_orders_admin(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    status: Optional[str] = None,
    payment_status: Optional[str] = None,
    cart_type: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    query = select(Order).options(selectinload(Order.items), selectinload(Order.user))
    if status:
        try:
            os_ = OrderStatus(status)
            query = query.where(Order.status == os_)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid status: {status}")
    if payment_status:
        try:
            ps_ = PaymentStatus(payment_status)
            query = query.where(Order.payment_status == ps_)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid payment_status: {payment_status}")
    if cart_type:
        query = query.where(Order.cart_type == cart_type)
    if search:
        query = query.join(User, Order.user_id == User.id).where(
            User.name.ilike(f"%{search}%") |
            User.email.ilike(f"%{search}%") |
            Order.id.ilike(f"%{search}%")
        )
    query = query.order_by(Order.created_at.desc()).offset((page - 1) * limit).limit(limit)
    result = await db.execute(query)
    orders = result.scalars().all()

    count_query = select(func.count(Order.id))
    if status:
        count_query = count_query.where(Order.status == OrderStatus(status))
    if payment_status:
        count_query = count_query.where(Order.payment_status == PaymentStatus(payment_status))
    if cart_type:
        count_query = count_query.where(Order.cart_type == cart_type)
    total = (await db.execute(count_query)).scalar_one()

    def order_dict(o: Order):
        d = o.to_dict()
        d["user_name"] = o.user.name if o.user else None
        d["user_email"] = o.user.email if o.user else None
        d["user_phone"] = o.user.phone if o.user else None
        return d

    return {"orders": [order_dict(o) for o in orders], "total": total, "page": page, "limit": limit}


@router.get("/orders/export")
async def export_orders_excel(
    date_from: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="End date YYYY-MM-DD"),
    status: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    """Export orders as an Excel (.xlsx) file for a given date range."""
    import io
    from datetime import date
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    query = select(Order).options(selectinload(Order.items), selectinload(Order.user))

    if date_from:
        try:
            df = datetime.strptime(date_from, "%Y-%m-%d")
            query = query.where(Order.created_at >= df)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_from format, use YYYY-MM-DD")

    if date_to:
        try:
            dt = datetime.strptime(date_to, "%Y-%m-%d")
            # Include the full day_to
            from datetime import timedelta
            dt = dt + timedelta(days=1)
            query = query.where(Order.created_at < dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_to format, use YYYY-MM-DD")

    if status:
        try:
            query = query.where(Order.status == OrderStatus(status))
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid status: {status}")

    query = query.order_by(Order.created_at.desc())
    result = await db.execute(query)
    orders = result.scalars().all()

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")

    headers = [
        "Order ID", "Date", "Customer Name", "Customer Email",
        "Status", "Payment Status", "Cart Type",
        "Shipping Type", "Items Count", "Subtotal",
        "Shipping Fee", "Commission", "Discount", "Total", "Currency"
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for order in orders:
        items_count = sum(i.quantity for i in order.items) if order.items else 0
        items_total = sum(i.price * i.quantity for i in order.items) if order.items else 0.0
        shipping_fee = 0.0
        commission = 0.0
        discount = order.discount_amount or 0.0
        total = order.total or 0.0
        # Derive shipping/commission from total breakdown
        shipping_fee = round(total - items_total + discount, 2) if total else 0.0

        ws.append([
            order.id[:8].upper(),
            order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else "",
            order.user.name if order.user else "",
            order.user.email if order.user else "",
            order.status.value if order.status else "",
            order.payment_status.value if order.payment_status else "",
            order.cart_type or "",
            order.shipping_type or "",
            items_count,
            round(items_total, 2),
            shipping_fee,
            commission,
            discount,
            total,
            order.currency or "SAR",
        ])

    # Auto-fit columns
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=row_idx, column=col_idx).value or ""))
            for row_idx in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"orders_{date_from or 'all'}_{date_to or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Pricing Policy ────────────────────────────────────────────────────────────

PRICING_POLICY_KEY = "pricing_policy"


class PricingPolicyRequest(BaseModel):
    shipping_mode: str  # "fixed" | "formula"
    shipping_value: float  # fixed amount OR multiplier (e.g. 0.05)
    shipping_hidden: bool  # hide shipping statement on product page
    commission_mode: str  # "fixed" | "formula"
    commission_value: float
    commission_hidden: bool


@router.get("/pricing-policy")
async def get_pricing_policy(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    """Get global pricing policy (shipping & commission defaults)."""
    import json
    from app.models.app_setting import AppSetting
    result = await db.execute(select(AppSetting).where(AppSetting.key == PRICING_POLICY_KEY))
    setting = result.scalar_one_or_none()
    if not setting:
        return {
            "shipping_mode": "fixed",
            "shipping_value": 0.0,
            "shipping_hidden": False,
            "commission_mode": "fixed",
            "commission_value": 0.0,
            "commission_hidden": False,
        }
    try:
        return json.loads(setting.value_en)
    except Exception:
        return json.loads("{}")


@router.post("/pricing-policy")
async def save_pricing_policy(
    req: PricingPolicyRequest,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    """Save global pricing policy (shipping & commission defaults)."""
    import json
    from app.models.app_setting import AppSetting
    payload = json.dumps(req.dict())
    result = await db.execute(select(AppSetting).where(AppSetting.key == PRICING_POLICY_KEY))
    setting = result.scalar_one_or_none()
    if not setting:
        setting = AppSetting(key=PRICING_POLICY_KEY, value_en=payload, value_ar=payload)
        db.add(setting)
    else:
        setting.value_en = payload
        setting.value_ar = payload
    await db.commit()
    return req.dict()


# ── Payment Approval ──────────────────────────────────────────────────────────

class PaymentActionRequest(BaseModel):
    reason: Optional[str] = None

@router.get("/payments/pending")
async def list_pending_payments(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    """List all orders awaiting manual payment approval."""
    result = await db.execute(
        select(Order)
        .where(Order.payment_status == PaymentStatus.PENDING_APPROVAL)
        .options(selectinload(Order.items), selectinload(Order.user))
        .order_by(Order.created_at.asc())
    )
    orders = result.scalars().all()

    def _fmt(o: Order):
        d = o.to_dict()
        d["user_name"] = o.user.name if o.user else None
        d["user_email"] = o.user.email if o.user else None
        return d

    return [_fmt(o) for o in orders]


@router.post("/payments/{order_id}/approve")
async def approve_payment(
    order_id: str,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    """Approve a manual payment and confirm the order."""
    result = await db.execute(
        select(Order).where(Order.id == order_id).options(selectinload(Order.user))
    )
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.payment_status != PaymentStatus.PENDING_APPROVAL:
        raise HTTPException(status_code=400, detail="Order is not pending payment approval")

    order.payment_status = PaymentStatus.APPROVED
    order.status = OrderStatus.CONFIRMED

    notif = Notification(
        user_id=order.user_id,
        title="Payment Approved ✅",
        message=f"Your payment for order #{order.id[:8].upper()} has been approved. Your order is confirmed!",
        type="order_status",
    )
    db.add(notif)
    await db.commit()

    if order.user:
        from app.email_service import send_payment_approved_email
        background_tasks.add_task(
            send_payment_approved_email,
            user_email=order.user.email,
            user_name=order.user.name or order.user.email,
            order_id=order.id,
            total=order.total,
        )

    return {"message": "Payment approved", "order_id": order.id}


@router.post("/payments/{order_id}/reject")
async def reject_payment(
    order_id: str,
    req: PaymentActionRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    """Reject a manual payment, cancel the order. Wallet payments get refunded."""
    result = await db.execute(
        select(Order).where(Order.id == order_id).options(selectinload(Order.user))
    )
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.payment_status not in (PaymentStatus.PENDING_APPROVAL, PaymentStatus.NOT_REQUIRED):
        raise HTTPException(status_code=400, detail="Order is not eligible for rejection")

    order.payment_status = PaymentStatus.REJECTED
    order.status = OrderStatus.CANCELLED

    # Refund wallet if original payment was wallet-based
    if order.payment_method_id == "wallet" and order.user:
        order.user.credit_balance = round((order.user.credit_balance or 0) + order.total, 2)
        tx = WalletTransaction(
            user_id=order.user_id,
            amount=order.total,
            type=WalletTransactionType.CREDIT,
            reason=f"Refund for rejected order #{order.id[:8].upper()}",
            reference_id=order.id,
            reference_type="order",
            balance_after=order.user.credit_balance,
        )
        db.add(tx)

    notif = Notification(
        user_id=order.user_id,
        title="Payment Rejected ❌",
        message=f"Your payment for order #{order.id[:8].upper()} was rejected and the order has been cancelled.",
        type="order_status",
    )
    db.add(notif)
    await db.commit()

    if order.user:
        from app.email_service import send_payment_rejected_email
        background_tasks.add_task(
            send_payment_rejected_email,
            user_email=order.user.email,
            user_name=order.user.name or order.user.email,
            order_id=order.id,
            total=order.total,
            reason=req.reason,
        )

    return {"message": "Payment rejected", "order_id": order.id}


# ── Admin Contact User ────────────────────────────────────────────────────────

class ContactUserRequest(BaseModel):
    message: str

@router.post("/orders/{order_id}/contact-user")
async def contact_user(
    order_id: str,
    req: ContactUserRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_admin_user),
):
    """Admin sends a message to the user regarding an order. Creates a support ticket."""
    order_result = await db.execute(
        select(Order).where(Order.id == order_id).options(selectinload(Order.user))
    )
    order = order_result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    user = order.user
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Create support ticket
    ticket = SupportTicket(
        user_id=user.id,
        title=f"Message from Support — Order #{order_id[:8].upper()}",
        description=req.message,
        status="open",
        user_unread=True,
    )
    db.add(ticket)
    await db.flush()

    # Add first message from admin
    msg = SupportMessage(
        ticket_id=ticket.id,
        sender="admin",
        message=req.message,
    )
    db.add(msg)

    # Notification for user
    notif = Notification(
        user_id=user.id,
        title="Message from Support 💬",
        message=f"You have a new message regarding order #{order_id[:8].upper()}.",
        type="support",
    )
    db.add(notif)
    await db.commit()

    from app.email_service import send_admin_contact_to_user_email
    background_tasks.add_task(
        send_admin_contact_to_user_email,
        user_email=user.email,
        user_name=user.name or user.email,
        order_id=order_id,
        admin_message=req.message,
    )

    return {"message": "Message sent", "ticket_id": ticket.id}


# ── Refund Management (Admin) ─────────────────────────────────────────────────

class RefundActionRequest(BaseModel):
    admin_note: Optional[str] = None

@router.get("/refunds")
async def list_refunds_admin(
    status: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    query = select(RefundRequest).options(
        selectinload(RefundRequest.user),
        selectinload(RefundRequest.order),
    )
    if status:
        try:
            rs = RefundStatus(status)
            query = query.where(RefundRequest.status == rs)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid status: {status}")
    query = query.order_by(RefundRequest.created_at.desc()).offset((page - 1) * limit).limit(limit)
    result = await db.execute(query)
    refunds = result.scalars().all()

    count_q = select(func.count(RefundRequest.id))
    if status:
        count_q = count_q.where(RefundRequest.status == RefundStatus(status))
    total = (await db.execute(count_q)).scalar_one()

    def _rfmt(r: RefundRequest):
        d = r.to_dict()
        d["user_name"] = r.user.name if r.user else None
        d["user_email"] = r.user.email if r.user else None
        d["order_total"] = r.order.total if r.order else None
        d["order_cart_type"] = r.order.cart_type if r.order else None
        return d

    return {"refunds": [_rfmt(r) for r in refunds], "total": total, "page": page, "limit": limit}


@router.post("/refunds/{refund_id}/approve")
async def approve_refund(
    refund_id: str,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    """Approve a refund: credit user wallet, log transaction, notify user."""
    result = await db.execute(
        select(RefundRequest).where(RefundRequest.id == refund_id)
        .options(selectinload(RefundRequest.user), selectinload(RefundRequest.order))
    )
    refund = result.scalar_one_or_none()
    if not refund:
        raise HTTPException(status_code=404, detail="Refund request not found")
    if refund.status != RefundStatus.PENDING:
        raise HTTPException(status_code=400, detail="Refund is not in pending status")

    refund.status = RefundStatus.APPROVED
    user = refund.user
    order = refund.order
    refund_amount = order.total if order else 0.0

    # Credit wallet
    if user:
        user.credit_balance = round((user.credit_balance or 0) + refund_amount, 2)
        tx = WalletTransaction(
            user_id=user.id,
            amount=refund_amount,
            type=WalletTransactionType.CREDIT,
            reason=f"Refund approved for order #{order.id[:8].upper() if order else 'N/A'}",
            reference_id=refund.id,
            reference_type="refund",
            balance_after=user.credit_balance,
        )
        db.add(tx)

    notif = Notification(
        user_id=refund.user_id,
        title="Refund Approved ✅",
        message=f"Your refund of {refund_amount:.2f} SAR has been approved and added to your wallet.",
        type="order_status",
    )
    db.add(notif)
    await db.commit()

    if user and order:
        from app.email_service import send_refund_status_email
        background_tasks.add_task(
            send_refund_status_email,
            user_email=user.email,
            user_name=user.name or user.email,
            order_id=order.id,
            amount=refund_amount,
            approved=True,
        )

    return {"message": "Refund approved", "refund_id": refund_id}


@router.post("/refunds/{refund_id}/reject")
async def reject_refund(
    refund_id: str,
    req: RefundActionRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    """Reject a refund request with an optional admin note."""
    result = await db.execute(
        select(RefundRequest).where(RefundRequest.id == refund_id)
        .options(selectinload(RefundRequest.user), selectinload(RefundRequest.order))
    )
    refund = result.scalar_one_or_none()
    if not refund:
        raise HTTPException(status_code=404, detail="Refund request not found")
    if refund.status != RefundStatus.PENDING:
        raise HTTPException(status_code=400, detail="Refund is not in pending status")

    refund.status = RefundStatus.REJECTED
    refund.admin_note = req.admin_note

    notif = Notification(
        user_id=refund.user_id,
        title="Refund Rejected ❌",
        message=f"Your refund request has been rejected.{' Reason: ' + req.admin_note if req.admin_note else ''}",
        type="order_status",
    )
    db.add(notif)
    await db.commit()

    user = refund.user
    order = refund.order
    if user and order:
        from app.email_service import send_refund_status_email
        background_tasks.add_task(
            send_refund_status_email,
            user_email=user.email,
            user_name=user.name or user.email,
            order_id=order.id,
            amount=order.total,
            approved=False,
            admin_note=req.admin_note,
        )

    return {"message": "Refund rejected", "refund_id": refund_id}


# ── Wallet Management (Admin) ─────────────────────────────────────────────────

class WalletAdjustRequest(BaseModel):
    amount: float
    type: str  # "credit" or "debit"
    reason: str

@router.post("/users/{user_id}/wallet-adjust")
async def adjust_wallet(
    user_id: str,
    req: WalletAdjustRequest,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_admin_user),
):
    """Admin manually credits or debits a user's wallet balance."""
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if req.type not in ("credit", "debit"):
        raise HTTPException(status_code=400, detail="type must be 'credit' or 'debit'")
    if req.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be positive")
    if req.type == "debit" and user.credit_balance < req.amount:
        raise HTTPException(status_code=400, detail="Insufficient wallet balance")

    if req.type == "credit":
        user.credit_balance = round(user.credit_balance + req.amount, 2)
        tx_type = WalletTransactionType.CREDIT
    else:
        user.credit_balance = round(user.credit_balance - req.amount, 2)
        tx_type = WalletTransactionType.DEBIT

    tx = WalletTransaction(
        user_id=user.id,
        amount=req.amount,
        type=tx_type,
        reason=f"[Admin Adjustment] {req.reason}",
        reference_type="admin_adjustment",
        balance_after=user.credit_balance,
    )
    db.add(tx)

    # Notification for user
    notif = Notification(
        user_id=user.id,
        title="Wallet Updated",
        message=f"Your wallet has been {'credited' if req.type == 'credit' else 'debited'} {req.amount:.2f} SAR. Reason: {req.reason}",
        type="order_status",
    )
    db.add(notif)
    await db.commit()
    return user.to_dict()


@router.get("/users/{user_id}/wallet-transactions")
async def get_wallet_transactions(
    user_id: str,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    """Get full wallet transaction ledger for a user."""
    result_u = await db.execute(select(User).where(User.id == user_id))
    if not result_u.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="User not found")

    result = await db.execute(
        select(WalletTransaction)
        .where(WalletTransaction.user_id == user_id)
        .order_by(WalletTransaction.created_at.desc())
        .offset((page - 1) * limit).limit(limit)
    )
    txs = result.scalars().all()
    total = (await db.execute(
        select(func.count(WalletTransaction.id)).where(WalletTransaction.user_id == user_id)
    )).scalar_one()
    return {"transactions": [t.to_dict() for t in txs], "total": total, "page": page, "limit": limit}


class UpdateOrderStatusRequest(BaseModel):
    status: str


@router.patch("/orders/{order_id}/status")
async def update_order_status(
    order_id: str,
    req: UpdateOrderStatusRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    result = await db.execute(
        select(Order).where(Order.id == order_id).options(selectinload(Order.user))
    )
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    try:
        order.status = OrderStatus(req.status)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid status: {req.status}")

    notif = Notification(
        user_id=order.user_id,
        title="Order Status Updated",
        message=f"Your order #{order.id[:8].upper()} status has been updated to {order.status.value}.",
        type="order_status"
    )
    db.add(notif)
    await db.commit()
    return {"message": "Status updated", "status": order.status.value}


# ── Categories ────────────────────────────────────────────────────────────────

@router.get("/categories")
async def list_categories_admin(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    result = await db.execute(select(Category).order_by(Category.sort_order))
    categories = result.scalars().all()
    return [c.to_dict() for c in categories]


class CategoryRequest(BaseModel):
    name_en: str
    name_ar: str
    icon: Optional[str] = None
    image_url: Optional[str] = None
    sort_order: int = 0


@router.post("/categories")
async def create_category(
    req: CategoryRequest,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    import uuid
    cat = Category(id=str(uuid.uuid4()), **req.model_dump())
    db.add(cat)
    await db.commit()
    await db.refresh(cat)
    return cat.to_dict()


@router.patch("/categories/{cat_id}")
async def update_category(
    cat_id: str,
    req: CategoryRequest,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    result = await db.execute(select(Category).where(Category.id == cat_id))
    cat = result.scalar_one_or_none()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    for field, value in req.model_dump(exclude_none=True).items():
        setattr(cat, field, value)
    await db.commit()
    return cat.to_dict()


@router.delete("/categories/{cat_id}")
async def delete_category(
    cat_id: str,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user),
):
    result = await db.execute(select(Category).where(Category.id == cat_id))
    cat = result.scalar_one_or_none()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    await db.delete(cat)
    await db.commit()
    return {"message": "Category deleted"}


# ── Admin login (creates/finds admin user) ────────────────────────────────────

class AdminLoginRequest(BaseModel):
    email: str
    password: str


@router.post("/login")
async def admin_login(req: AdminLoginRequest, db: AsyncSession = Depends(get_db)):
    """Admin-specific login endpoint that only accepts admin-role users."""
    from app.auth.security import verify_password
    from app.auth.jwt_handler import create_access_token, create_refresh_token

    result = await db.execute(select(User).where(User.email == req.email, User.role == "admin"))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials or not an admin")
    if not user.is_active:
        raise HTTPException(status_code=401, detail="Admin account is inactive")
    if not user.password_hash:
        raise HTTPException(status_code=401, detail="Password login not configured for this account")
    if not verify_password(req.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    access_token = create_access_token({"sub": user.id, "role": user.role})
    refresh_token = create_refresh_token({"sub": user.id})
    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "user": user.to_dict(),
    }


@router.post("/seed-admin")
async def seed_admin(db: AsyncSession = Depends(get_db)):
    """One-time endpoint to create the default admin account. Remove in production."""
    from app.auth.security import hash_password

    result = await db.execute(select(User).where(User.role == "admin"))
    existing = result.scalar_one_or_none()
    if existing:
        if not existing.is_active:
            existing.is_active = True
            await db.commit()
            return {"message": "Admin account activated", "email": existing.email}
        return {"message": "Admin already exists", "email": existing.email}

    password_hash = hash_password("admin123")
    admin = User(
        email="admin@widdistore.com",
        password_hash=password_hash,
        name="Widdi Admin",
        role="admin",
        is_active=True,
        is_verified=True,
    )
    db.add(admin)
    await db.commit()
    return {"message": "Admin created", "email": "admin@widdistore.com", "password": "admin123"}


# ── States & Cities ──────────────────────────────────────────────────────────

class StateRequest(BaseModel):
    name_en: str
    name_ar: str
    shipping_fee: Optional[float] = 0.0
    commission: Optional[float] = 0.0
    free_shipping: Optional[bool] = False
    no_commission: Optional[bool] = False

class CityRequest(BaseModel):
    state_id: str
    name_en: str
    name_ar: str
    shipping_fee: Optional[float] = 0.0
    commission: Optional[float] = 0.0
    free_shipping: Optional[bool] = False
    no_commission: Optional[bool] = False

@router.get("/states")
async def list_states_admin(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    result = await db.execute(select(State).order_by(State.name_en))
    return [s.to_dict() for s in result.scalars().all()]

@router.post("/states")
async def create_state_admin(
    req: StateRequest,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    state = State(**req.model_dump())
    db.add(state)
    await db.commit()
    await db.refresh(state)
    return state.to_dict()

@router.patch("/states/{state_id}")
async def update_state_admin(
    state_id: str,
    req: StateRequest,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    result = await db.execute(select(State).where(State.id == state_id))
    state = result.scalar_one_or_none()
    if not state:
        raise HTTPException(status_code=404, detail="State not found")
    state.name_en = req.name_en
    state.name_ar = req.name_ar
    state.shipping_fee = req.shipping_fee if req.shipping_fee is not None else 0.0
    state.commission = req.commission if req.commission is not None else 0.0
    state.free_shipping = req.free_shipping if req.free_shipping is not None else False
    state.no_commission = req.no_commission if req.no_commission is not None else False
    await db.commit()
    await db.refresh(state)
    return state.to_dict()

@router.delete("/states/{state_id}")
async def delete_state_admin(
    state_id: str,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    result = await db.execute(select(State).where(State.id == state_id))
    state = result.scalar_one_or_none()
    if not state:
        raise HTTPException(status_code=404, detail="State not found")
    await db.delete(state)
    await db.commit()
    return {"message": "State deleted"}

@router.get("/cities")
async def list_cities_admin(
    state_id: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    query = select(City)
    if state_id:
        query = query.where(City.state_id == state_id)
    query = query.order_by(City.name_en)
    result = await db.execute(query)
    return [c.to_dict() for c in result.scalars().all()]

@router.post("/cities")
async def create_city_admin(
    req: CityRequest,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    # Verify state exists
    state_check = await db.execute(select(State).where(State.id == req.state_id))
    if not state_check.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="State not found")
    city = City(**req.model_dump())
    db.add(city)
    await db.commit()
    await db.refresh(city)
    return city.to_dict()

@router.patch("/cities/{city_id}")
async def update_city_admin(
    city_id: str,
    req: CityRequest,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    result = await db.execute(select(City).where(City.id == city_id))
    city = result.scalar_one_or_none()
    if not city:
        raise HTTPException(status_code=404, detail="City not found")
    
    state_check = await db.execute(select(State).where(State.id == req.state_id))
    if not state_check.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="State not found")

    city.state_id = req.state_id
    city.name_en = req.name_en
    city.name_ar = req.name_ar
    city.shipping_fee = req.shipping_fee if req.shipping_fee is not None else 0.0
    city.commission = req.commission if req.commission is not None else 0.0
    city.free_shipping = req.free_shipping if req.free_shipping is not None else False
    city.no_commission = req.no_commission if req.no_commission is not None else False
    await db.commit()
    await db.refresh(city)
    return city.to_dict()

@router.delete("/cities/{city_id}")
async def delete_city_admin(
    city_id: str,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    result = await db.execute(select(City).where(City.id == city_id))
    city = result.scalar_one_or_none()
    if not city:
        raise HTTPException(status_code=404, detail="City not found")
    await db.delete(city)
    await db.commit()
    return {"message": "City deleted"}


# ── Payment Methods ──────────────────────────────────────────────────────────

class PaymentMethodRequest(BaseModel):
    title_en: str
    title_ar: str
    description_en: Optional[str] = None
    description_ar: Optional[str] = None
    details_en: Optional[str] = None
    details_ar: Optional[str] = None
    image_url: Optional[str] = None
    is_active: bool = True
    fields: Optional[list] = None  # list of field dicts: [{"key": "...", "label_en": "...", "label_ar": "..."}]
    bank_accounts: Optional[list] = None  # list of bank account dicts: [{"bank_name_en": "...", "bank_name_ar": "...", "account_number": "...", "logo_url": "..."}]


@router.get("/payment-methods")
async def list_payment_methods_admin(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    from app.models.payment_method import PaymentMethod
    result = await db.execute(select(PaymentMethod))
    methods = result.scalars().all()
    return [m.to_dict("en") for m in methods]


@router.post("/payment-methods")
async def create_payment_method_admin(
    req: PaymentMethodRequest,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    from app.models.payment_method import PaymentMethod
    import json
    import uuid
    fields_str = json.dumps(req.fields or [])
    bank_accounts_str = json.dumps(req.bank_accounts or [])
    method = PaymentMethod(
        id=str(uuid.uuid4()),
        title_en=req.title_en,
        title_ar=req.title_ar,
        description_en=req.description_en,
        description_ar=req.description_ar,
        details_en=req.details_en,
        details_ar=req.details_ar,
        image_url=req.image_url,
        is_active=req.is_active,
        fields_json=fields_str,
        bank_accounts_json=bank_accounts_str
    )
    db.add(method)
    await db.commit()
    await db.refresh(method)
    return method.to_dict("en")


@router.patch("/payment-methods/{method_id}")
async def update_payment_method_admin(
    method_id: str,
    req: PaymentMethodRequest,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    from app.models.payment_method import PaymentMethod
    import json
    result = await db.execute(select(PaymentMethod).where(PaymentMethod.id == method_id))
    method = result.scalar_one_or_none()
    if not method:
        raise HTTPException(status_code=404, detail="Payment method not found")
    
    method.title_en = req.title_en
    method.title_ar = req.title_ar
    method.description_en = req.description_en
    method.description_ar = req.description_ar
    method.details_en = req.details_en
    method.details_ar = req.details_ar
    method.image_url = req.image_url
    method.is_active = req.is_active
    if req.fields is not None:
        method.fields_json = json.dumps(req.fields)
    if req.bank_accounts is not None:
        method.bank_accounts_json = json.dumps(req.bank_accounts)
        
    await db.commit()
    await db.refresh(method)
    return method.to_dict("en")


@router.delete("/payment-methods/{method_id}")
async def delete_payment_method_admin(
    method_id: str,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    from app.models.payment_method import PaymentMethod
    result = await db.execute(select(PaymentMethod).where(PaymentMethod.id == method_id))
    method = result.scalar_one_or_none()
    if not method:
        raise HTTPException(status_code=404, detail="Payment method not found")
    await db.delete(method)
    await db.commit()
    return {"message": "Payment method deleted"}


# ── Coupon Administration Endpoints ────────────────────────────────────────

class CouponRequest(BaseModel):
    code: str
    description_en: Optional[str] = None
    description_ar: Optional[str] = None
    discount_type: str = "percentage"  # percentage, fixed
    discount_value: float
    min_order_amount: float = 0.0
    max_discount: Optional[float] = None
    usage_limit: Optional[int] = None
    is_active: bool = True
    expires_at: Optional[str] = None  # ISO format string
    applicability: str = "all"  # all, internal, external


@router.get("/coupons")
async def list_coupons_admin(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    # Total count
    count_query = select(func.count(Coupon.id))
    result_count = await db.execute(count_query)
    total = result_count.scalar() or 0

    query = select(Coupon).order_by(Coupon.created_at.desc()).offset((page - 1) * limit).limit(limit)
    result = await db.execute(query)
    coupons = result.scalars().all()
    return {
        "total": total,
        "coupons": [c.to_dict("en") for c in coupons]
    }


@router.post("/coupons")
async def create_coupon_admin(
    req: CouponRequest,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    # Check uniqueness
    existing = await db.execute(select(Coupon).where(Coupon.code == req.code))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Coupon code already exists")

    expires = None
    if req.expires_at:
        try:
            expires = datetime.fromisoformat(req.expires_at.replace("Z", "+00:00")).replace(tzinfo=None)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid date format for expires_at")

    coupon = Coupon(
        code=req.code.upper().strip(),
        description_en=req.description_en,
        description_ar=req.description_ar,
        discount_type=req.discount_type,
        discount_value=req.discount_value,
        min_order_amount=req.min_order_amount,
        max_discount=req.max_discount,
        usage_limit=req.usage_limit,
        is_active=req.is_active,
        expires_at=expires,
        applicability=req.applicability,
    )
    db.add(coupon)
    await db.commit()
    await db.refresh(coupon)
    return coupon.to_dict("en")


@router.put("/coupons/{coupon_id}")
async def update_coupon_admin(
    coupon_id: str,
    req: CouponRequest,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    result = await db.execute(select(Coupon).where(Coupon.id == coupon_id))
    coupon = result.scalar_one_or_none()
    if not coupon:
        raise HTTPException(status_code=404, detail="Coupon not found")

    # Check uniqueness if code is changed
    if coupon.code != req.code.upper().strip():
        existing = await db.execute(select(Coupon).where(Coupon.code == req.code.upper().strip()))
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Coupon code already exists")

    expires = None
    if req.expires_at:
        try:
            expires = datetime.fromisoformat(req.expires_at.replace("Z", "+00:00")).replace(tzinfo=None)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid date format for expires_at")

    coupon.code = req.code.upper().strip()
    coupon.description_en = req.description_en
    coupon.description_ar = req.description_ar
    coupon.discount_type = req.discount_type
    coupon.discount_value = req.discount_value
    coupon.min_order_amount = req.min_order_amount
    coupon.max_discount = req.max_discount
    coupon.usage_limit = req.usage_limit
    coupon.is_active = req.is_active
    coupon.expires_at = expires
    coupon.applicability = req.applicability

    await db.commit()
    await db.refresh(coupon)
    return coupon.to_dict("en")


@router.delete("/coupons/{coupon_id}")
async def delete_coupon_admin(
    coupon_id: str,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    result = await db.execute(select(Coupon).where(Coupon.id == coupon_id))
    coupon = result.scalar_one_or_none()
    if not coupon:
        raise HTTPException(status_code=404, detail="Coupon not found")

    await db.delete(coupon)
    await db.commit()
    return {"message": "Coupon deleted"}


# ── App Settings ─────────────────────────────────────────────────────────────

class AppSettingUpdateRequest(BaseModel):
    value_en: str
    value_ar: str


@router.get("/settings")
async def list_settings_admin(
    db: AsyncSession = Depends(get_db)
):
    from app.models.app_setting import AppSetting
    result = await db.execute(select(AppSetting))
    settings = result.scalars().all()
    return {s.key: s.to_dict() for s in settings}


@router.put("/settings/{key}")
async def update_setting_admin(
    key: str,
    req: AppSettingUpdateRequest,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    from app.models.app_setting import AppSetting
    result = await db.execute(select(AppSetting).where(AppSetting.key == key))
    setting = result.scalar_one_or_none()
    if not setting:
        setting = AppSetting(key=key, value_en=req.value_en, value_ar=req.value_ar)
        db.add(setting)
    else:
        setting.value_en = req.value_en
        setting.value_ar = req.value_ar
    await db.commit()
    await db.refresh(setting)
    return setting.to_dict()


